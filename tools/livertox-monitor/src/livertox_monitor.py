#!/usr/bin/env python3
"""LiverTox source monitor — detects what changed between quarterly runs.

Implements DEC-P34 (source freshness monitoring), programmatic sources P1-P3.

This is admin tooling. It fetches three public LiverTox endpoints, stores a
snapshot, and writes a human-readable change report. It reports ONLY: it never
writes to a substance table, never assigns any classification of its own, and
never interprets a Likelihood Score. Every changed or added entry is labeled for
human disposition.

Usage:
    python livertox_monitor.py --baseline    # first run: snapshot + counts
    python livertox_monitor.py --check       # default: diff, report, snapshot

Exit codes:
    0  ran successfully (whether or not changes were found)
    1  network, parse, or structure failure — the prior snapshot is left intact
"""

# ---------------------------------------------------------------------------
# WHY THIS SCRIPT IS DELIBERATELY DUMB
#
# Per DEC-P33, the LiverTox Likelihood Score measures hepatotoxicity only. It is
# systematically wrong as a general transplant risk signal: acetaminophen scores
# A and is the recommended post-transplant analgesic, while pseudoephedrine
# scores E and transplant centers say to avoid it. Any code here that compared
# scores for severity, ranked them, or mapped them onto a tier would be encoding
# a clinical judgment this script has no business making.
#
# So score strings are carried through as OPAQUE TEXT. They are compared for
# equality (changed / not changed) and never for magnitude. 'A [HD]', 'C[HD]'
# and 'A ' are preserved verbatim, including their inconsistent spacing.
#
# Per DEC-P34 rule 1, a score downgrade is a review trigger, never a deletion:
# this script reports the change and never suggests removal.
# ---------------------------------------------------------------------------

from __future__ import annotations

import argparse
import hashlib
import io
import json
import re
import sys
from datetime import date, datetime
from html import unescape
from pathlib import Path
from urllib.parse import urljoin

try:
    import requests
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - environment problem, not logic
    sys.stderr.write(f"Missing dependency: {exc}. Run: pip install -r requirements.txt\n")
    raise SystemExit(1)

# ── Endpoints (DEC-P34 Tier 1) ─────────────────────────────────────────────
CHANGELOG_URL = "https://www.ncbi.nlm.nih.gov/books/n/livertox/updates/"
MASTERLIST_PAGE_URL = "https://www.ncbi.nlm.nih.gov/books/n/livertox/masterlistintro/"
HDS_INDEX_URL = "https://www.ncbi.nlm.nih.gov/books/n/livertox/HerbalDietarySuppl/"

USER_AGENT = (
    "InsinaHealth-SourceMonitor/1.0 (admin tooling; quarterly source freshness check; "
    "contact: gregbutler205@gmail.com)"
)
TIMEOUT = 60

# Columns the diff cannot work without. Located BY HEADER TEXT, never by
# position — column order is not guaranteed across releases.
REQUIRED_HEADERS = ["Ingredient", "Brand Name", "Likelihood Score", "Primary Classification"]
OPTIONAL_HEADERS = ["Count", "Chapter Title", "Last Update", "Year Approved",
                    "In LiverTox", "Secondary Classification"]

HOLDING_QUEUE = "HOLDING QUEUE — requires human disposition"

ROOT = Path(__file__).resolve().parent.parent
SNAPSHOT_DIR = ROOT / "snapshots"
REPORT_DIR = ROOT / "reports"


def _force_utf8_streams() -> None:
    """Print UTF-8 regardless of console/locale.

    The report is Markdown containing em dashes and arrows. On a Windows
    console (cp1252) those render as replacement characters, and when stdout is
    redirected to a file they raise UnicodeEncodeError outright. Files are
    already written with an explicit encoding; this makes the terminal and any
    redirect behave the same way.
    """
    for stream in (sys.stdout, sys.stderr):
        try:
            stream.reconfigure(encoding="utf-8")
        except (AttributeError, ValueError):  # pragma: no cover - exotic stream
            pass


_force_utf8_streams()


class SourceStructureError(RuntimeError):
    """The page or spreadsheet did not look the way this script requires."""


# ── Fetching ────────────────────────────────────────────────────────────────
def _session() -> requests.Session:
    s = requests.Session()
    s.headers.update({"User-Agent": USER_AGENT})
    return s


def fetch_text(url: str, session: requests.Session) -> str:
    resp = session.get(url, timeout=TIMEOUT)
    resp.raise_for_status()
    return resp.text


def fetch_bytes(url: str, session: requests.Session) -> bytes:
    resp = session.get(url, timeout=TIMEOUT)
    resp.raise_for_status()
    return resp.content


# ── P2: resolve the spreadsheet link (never hardcoded) ──────────────────────
def resolve_xlsx_url(page_html: str, page_url: str = MASTERLIST_PAGE_URL) -> str:
    """Find the .xlsx href on the master list page.

    The filename encodes a release date (observed: masterlist02-26.xlsx) and
    changes on release, so it is resolved on every run. If no link is present
    this raises: there is deliberately no cached fallback, because silently
    diffing last quarter's file against itself would look like "no changes".
    """
    hrefs = re.findall(r'href\s*=\s*["\']([^"\']+\.xlsx[^"\']*)["\']', page_html, re.IGNORECASE)
    if not hrefs:
        raise SourceStructureError(
            "No .xlsx link found on the master list page (%s).\n"
            "The page structure may have changed, or the spreadsheet may have been "
            "withdrawn. Refusing to fall back to a previously seen URL — a stale "
            "spreadsheet would diff clean and hide real changes.\n"
            "Fix: open the page, find the current spreadsheet link, and update the "
            "link-detection here if the markup changed." % page_url
        )
    return urljoin(page_url, hrefs[0])


# ── P1: changelog ───────────────────────────────────────────────────────────
_ENTRY_RE = re.compile(r'<li[^>]*class="[^"]*half_rhythm[^"]*"[^>]*>\s*<div>(.*?)</div>\s*</li>', re.S | re.I)
_LINK_TEXT_RE = re.compile(r"<a[^>]*>(.*?)</a>", re.S | re.I)
_TAG_RE = re.compile(r"<[^>]+>")
_DATE_RE = re.compile(r"(\d{1,2}\s+[A-Za-z]{3,}\s+\d{4})")
_PAGE_DATE_RE = re.compile(r"Updated:\s*([0-9]{1,2}\s+[A-Za-z]{3,}\s+[0-9]{4})", re.I)


def _strip_tags(fragment: str) -> str:
    return unescape(_TAG_RE.sub("", fragment)).strip()


def parse_changelog(html: str) -> dict:
    """Extract the page's own 'Updated:' date and its dated entries.

    Entries look like: <li class="half_rhythm"><div><a ...>Cinnamon</a> 01 Jul 2026</div></li>
    """
    body = html[html.find("body-content"):] or html
    page_date_match = _PAGE_DATE_RE.search(body)
    entries = []
    for fragment in _ENTRY_RE.findall(body):
        text = _strip_tags(fragment)
        if not text:
            continue
        link = _LINK_TEXT_RE.search(fragment)
        name = _strip_tags(link.group(1)) if link else text
        date_match = _DATE_RE.search(text)
        if not date_match:
            continue  # not a dated changelog row
        entries.append({
            "name": name,
            "date_text": date_match.group(1),
            "date_iso": _parse_day_month_year(date_match.group(1)),
        })
    if not entries:
        raise SourceStructureError(
            "No dated entries found on the changelog page (%s). The page markup "
            "likely changed; the changelog is the real currency signal (DEC-P34) "
            "so this is treated as a failure rather than 'nothing new'." % CHANGELOG_URL
        )
    return {
        "page_updated_text": page_date_match.group(1) if page_date_match else None,
        "page_updated_iso": _parse_day_month_year(page_date_match.group(1)) if page_date_match else None,
        "entries": entries,
    }


def _parse_day_month_year(text: str) -> str | None:
    """'01 Jul 2026' -> '2026-07-01'. Returns None if unparseable (kept as text)."""
    for fmt in ("%d %b %Y", "%d %B %Y"):
        try:
            return datetime.strptime(text.strip(), fmt).date().isoformat()
        except ValueError:
            continue
    return None


# ── P3: HDS index ───────────────────────────────────────────────────────────
_HDS_LINK_RE = re.compile(r'<a href="/books/n/livertox/([^/"]+)/"[^>]*>([^<]+)</a>', re.I)


def parse_hds_index(html: str) -> list[str]:
    body = html[html.find("body-content"):] or html
    names = {unescape(text).strip() for _slug, text in _HDS_LINK_RE.findall(body)}
    names.discard("")
    if not names:
        raise SourceStructureError(
            "No entries found on the HDS index page (%s); markup likely changed." % HDS_INDEX_URL
        )
    return sorted(names)


# ── P2: spreadsheet ─────────────────────────────────────────────────────────
def parse_masterlist(xlsx_bytes: bytes) -> dict:
    """Parse the master list, locating columns by header text.

    The sheet has a two-row header: row 1 is a title banner carrying the
    spreadsheet's own 'Last Update:' note, row 2 is the column headers, and data
    starts at row 3. Rather than trusting those positions, the header row is
    found by looking for the required header labels.
    """
    workbook = load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = [list(r) for r in sheet.iter_rows(values_only=True)]
    if not rows:
        raise SourceStructureError("The spreadsheet is empty.")

    header_index, columns = _locate_header_row(rows)
    banner = _find_banner_note(rows[:header_index])

    records = []
    for raw in rows[header_index + 1:]:
        ingredient = _cell(raw, columns.get("Ingredient"))
        if not ingredient:
            continue  # trailing blank rows
        record = {"Ingredient": ingredient}
        for label, idx in columns.items():
            if label == "Ingredient":
                continue
            record[label] = _cell(raw, idx)
        records.append(record)

    if not records:
        raise SourceStructureError("The spreadsheet had headers but no data rows.")

    return {
        "banner_note": banner,
        "columns_found": sorted(columns),
        "header_row_number": header_index + 1,
        "records": records,
    }


def _locate_header_row(rows: list[list]) -> tuple[int, dict[str, int]]:
    for index, row in enumerate(rows[:20]):  # header lives near the top
        labels = {}
        for position, value in enumerate(row):
            if value is None:
                continue
            text = str(value).strip()
            if text in REQUIRED_HEADERS or text in OPTIONAL_HEADERS:
                labels.setdefault(text, position)
        if all(h in labels for h in REQUIRED_HEADERS):
            return index, labels
    raise SourceStructureError(
        "Could not find a header row containing all required columns "
        f"({', '.join(REQUIRED_HEADERS)}). The spreadsheet layout changed; "
        "columns are located by header text on purpose, so this is a hard stop "
        "rather than a guess at positions."
    )


def _find_banner_note(rows_above_header: list[list]) -> str | None:
    for row in rows_above_header:
        for value in row:
            if value is None:
                continue
            text = str(value).strip()
            if text.lower().startswith("last update"):
                return text
    return None


def _cell(row: list, index: int | None) -> str:
    """Cell as text, verbatim apart from outer whitespace on non-score fields.

    Dates arrive as datetimes; they are rendered as ISO dates so a snapshot is
    JSON-serializable and stable across runs.
    """
    if index is None or index >= len(row):
        return ""
    value = row[index]
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


# ── Snapshot ────────────────────────────────────────────────────────────────
def build_snapshot(changelog: dict, hds_names: list[str], masterlist: dict,
                   xlsx_url: str, xlsx_sha256: str) -> dict:
    return {
        "schema": 1,
        "run_date": date.today().isoformat(),
        "generated_by": "livertox_monitor.py (DEC-P34)",
        "sources": {
            "changelog_url": CHANGELOG_URL,
            "masterlist_page_url": MASTERLIST_PAGE_URL,
            "hds_index_url": HDS_INDEX_URL,
            "xlsx_url": xlsx_url,
            "xlsx_sha256": xlsx_sha256,
        },
        "changelog": changelog,
        "hds_index": hds_names,
        "masterlist": masterlist,
    }


def snapshot_path(run_date: str) -> Path:
    """Date-stamped path that never collides with an existing snapshot."""
    SNAPSHOT_DIR.mkdir(parents=True, exist_ok=True)
    candidate = SNAPSHOT_DIR / f"livertox-{run_date}.json"
    counter = 2
    while candidate.exists():  # never overwrite a stored snapshot
        candidate = SNAPSHOT_DIR / f"livertox-{run_date}-{counter}.json"
        counter += 1
    return candidate


_SNAP_RE = re.compile(r"^livertox-(\d{4}-\d{2}-\d{2})(?:-(\d+))?\.json$")


def latest_snapshot() -> tuple[Path, dict] | tuple[None, None]:
    if not SNAPSHOT_DIR.exists():
        return None, None
    dated = []
    for path in SNAPSHOT_DIR.glob("livertox-*.json"):
        match = _SNAP_RE.match(path.name)
        if match:
            dated.append(((match.group(1), int(match.group(2) or 1)), path))
    if not dated:
        return None, None
    path = max(dated)[1]
    with path.open(encoding="utf-8") as handle:
        return path, json.load(handle)


# ── Diff ────────────────────────────────────────────────────────────────────
def _index_by_ingredient(records: list[dict]) -> dict[str, list[dict]]:
    """Group rows by ingredient name.

    Names are NOT unique in this source: 9 names appear twice in the observed
    release, sometimes with different scores (Omalizumab is listed as both 'X'
    and 'E' under different brands). Grouping keeps those honest instead of
    silently dropping one. Whitespace is normalized for the KEY only — 22 names
    carry trailing spaces — while the displayed value stays verbatim.
    """
    grouped: dict[str, list[dict]] = {}
    for record in records:
        key = record.get("Ingredient", "").strip().casefold()
        if key:
            grouped.setdefault(key, []).append(record)
    return grouped


def _score_list(records: list[dict]) -> list[str]:
    # Score strings are compared verbatim, for equality only — never ordered or
    # ranked. See the note at the top of this file.
    return sorted(r.get("Likelihood Score", "") for r in records)


def diff_snapshots(old: dict, new: dict) -> dict:
    old_rows = _index_by_ingredient(old["masterlist"]["records"])
    new_rows = _index_by_ingredient(new["masterlist"]["records"])

    added_keys = sorted(set(new_rows) - set(old_rows))
    removed_keys = sorted(set(old_rows) - set(new_rows))

    score_changes = []
    for key in sorted(set(old_rows) & set(new_rows)):
        before, after = _score_list(old_rows[key]), _score_list(new_rows[key])
        if before != after:
            score_changes.append({
                "ingredient": new_rows[key][0].get("Ingredient", "").strip(),
                "score_before": before,
                "score_after": after,
                "classification": new_rows[key][0].get("Primary Classification", ""),
            })

    old_hds, new_hds = set(old.get("hds_index", [])), set(new.get("hds_index", []))

    old_seen = {(e["name"], e["date_text"]) for e in old["changelog"]["entries"]}
    new_changelog = [e for e in new["changelog"]["entries"]
                     if (e["name"], e["date_text"]) not in old_seen]

    return {
        "masterlist_added": [_summarize(r) for key in added_keys for r in new_rows[key]],
        "masterlist_removed": [_summarize(r) for key in removed_keys for r in old_rows[key]],
        "masterlist_score_changed": score_changes,
        "hds_added": sorted(new_hds - old_hds),
        "hds_removed": sorted(old_hds - new_hds),
        "changelog_new": new_changelog,
    }


def _summarize(record: dict) -> dict:
    return {
        "ingredient": record.get("Ingredient", "").strip(),
        "brand": record.get("Brand Name", ""),
        "score": record.get("Likelihood Score", ""),
        "classification": record.get("Primary Classification", ""),
    }


def has_changes(changes: dict) -> bool:
    return any(changes[key] for key in changes)


# ── Report ──────────────────────────────────────────────────────────────────
def render_report(snapshot: dict, changes: dict | None, previous: dict | None) -> str:
    out: list[str] = []
    add = out.append
    masterlist = snapshot["masterlist"]

    add(f"# LiverTox source report — {snapshot['run_date']}")
    add("")
    add("Generated by `livertox_monitor.py` (DEC-P34). **Reports only.** Nothing here has "
        "been applied to any table, and no entry has been accepted, approved, or promoted.")
    add("")

    add("## Source dates (the lag is expected)")
    add("")
    add("| Source | Date it reports |")
    add("|---|---|")
    add(f"| Changelog page (P1) | {snapshot['changelog'].get('page_updated_text') or 'not stated'} |")
    add(f"| Spreadsheet banner (P2) | {masterlist.get('banner_note') or 'not stated'} |")
    add("")
    add("The spreadsheet trails the site. Per DEC-P34, read the changelog for what is "
        "actually new; the spreadsheet is the structured diff.")
    add("")

    add("## Snapshot metadata")
    add("")
    add(f"- Run date: `{snapshot['run_date']}`")
    add(f"- Resolved spreadsheet URL: `{snapshot['sources']['xlsx_url']}`")
    add(f"- Spreadsheet SHA-256: `{snapshot['sources']['xlsx_sha256']}`")
    add(f"- Spreadsheet rows: {len(masterlist['records'])}")
    add(f"- HDS index entries: {len(snapshot['hds_index'])}")
    add(f"- Changelog entries on page: {len(snapshot['changelog']['entries'])}")
    if previous:
        add(f"- Compared against snapshot: `{previous['run_date']}`")
    add("")

    if changes is None:
        add("## Baseline established")
        add("")
        add("No comparison was made — this run stored the first snapshot. Re-run with "
            "`--check` next quarter to see changes.")
        add("")
        return "\n".join(out)

    if not has_changes(changes):
        add("## No changes")
        add("")
        add("Nothing changed in the changelog, the spreadsheet, or the HDS index since the "
            "previous snapshot.")
        add("")
        return "\n".join(out)

    add("## Changes requiring disposition")
    add("")
    add(f"Every item below is **{HOLDING_QUEUE}**. Per DEC-P34, new and changed entries land "
        "in the holding queue, not production, and a score change is a review trigger — "
        "never an automatic removal.")
    add("")

    add(f"### Changelog entries new since the last snapshot ({len(changes['changelog_new'])})")
    add("")
    if changes["changelog_new"]:
        add("| Entry | Date | Disposition |")
        add("|---|---|---|")
        for entry in changes["changelog_new"]:
            add(f"| {entry['name']} | {entry['date_text']} | {HOLDING_QUEUE} |")
    else:
        add("_None._")
    add("")

    add(f"### Spreadsheet — Likelihood Score string changed ({len(changes['masterlist_score_changed'])})")
    add("")
    if changes["masterlist_score_changed"]:
        add("Scores are shown verbatim, exactly as published. They are not interpreted, "
            "ranked, or mapped to anything by this script.")
        add("")
        add("| Ingredient | Before | After | Primary classification | Disposition |")
        add("|---|---|---|---|---|")
        for change in changes["masterlist_score_changed"]:
            before = ", ".join(f"`{s}`" for s in change["score_before"])
            after = ", ".join(f"`{s}`" for s in change["score_after"])
            add(f"| {change['ingredient']} | {before} | {after} | "
                f"{change['classification']} | {HOLDING_QUEUE} |")
    else:
        add("_None._")
    add("")

    add(f"### Spreadsheet — ingredients added ({len(changes['masterlist_added'])})")
    add("")
    if changes["masterlist_added"]:
        add("| Ingredient | Brand | Score | Primary classification | Disposition |")
        add("|---|---|---|---|---|")
        for row in changes["masterlist_added"]:
            add(f"| {row['ingredient']} | {row['brand']} | `{row['score']}` | "
                f"{row['classification']} | {HOLDING_QUEUE} |")
    else:
        add("_None._")
    add("")

    add(f"### Spreadsheet — ingredients no longer listed ({len(changes['masterlist_removed'])})")
    add("")
    if changes["masterlist_removed"]:
        add("Absence from the spreadsheet is a fact about the source, not an instruction. "
            "It requires the same human disposition as any other change.")
        add("")
        add("| Ingredient | Brand | Last seen score | Primary classification | Disposition |")
        add("|---|---|---|---|---|")
        for row in changes["masterlist_removed"]:
            add(f"| {row['ingredient']} | {row['brand']} | `{row['score']}` | "
                f"{row['classification']} | {HOLDING_QUEUE} |")
    else:
        add("_None._")
    add("")

    add(f"### HDS index — entries added ({len(changes['hds_added'])})")
    add("")
    add(("\n".join(f"- {name} — {HOLDING_QUEUE}" for name in changes["hds_added"]))
        if changes["hds_added"] else "_None._")
    add("")

    add(f"### HDS index — entries removed ({len(changes['hds_removed'])})")
    add("")
    add(("\n".join(f"- {name} — {HOLDING_QUEUE}" for name in changes["hds_removed"]))
        if changes["hds_removed"] else "_None._")
    add("")
    return "\n".join(out)


# ── Main ────────────────────────────────────────────────────────────────────
def collect(session: requests.Session) -> tuple[dict, list[str], dict, str, str]:
    changelog = parse_changelog(fetch_text(CHANGELOG_URL, session))
    hds_names = parse_hds_index(fetch_text(HDS_INDEX_URL, session))
    xlsx_url = resolve_xlsx_url(fetch_text(MASTERLIST_PAGE_URL, session))
    xlsx_bytes = fetch_bytes(xlsx_url, session)
    masterlist = parse_masterlist(xlsx_bytes)
    return changelog, hds_names, masterlist, xlsx_url, hashlib.sha256(xlsx_bytes).hexdigest()


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Detect changes in LiverTox between quarterly runs (DEC-P34). Reports only."
    )
    mode = parser.add_mutually_exclusive_group()
    mode.add_argument("--baseline", action="store_true",
                      help="store the first snapshot and report counts")
    mode.add_argument("--check", action="store_true",
                      help="compare against the most recent snapshot (default)")
    args = parser.parse_args(argv)

    try:
        session = _session()
        changelog, hds_names, masterlist, xlsx_url, xlsx_sha = collect(session)
    except SourceStructureError as exc:
        sys.stderr.write(f"\nSOURCE STRUCTURE PROBLEM\n{exc}\n\nNo snapshot was written.\n")
        return 1
    except requests.RequestException as exc:
        sys.stderr.write(
            f"\nNETWORK FAILURE\nCould not reach a LiverTox endpoint: {exc}\n\n"
            "No snapshot was written; the previous snapshot is untouched.\n"
        )
        return 1

    snapshot = build_snapshot(changelog, hds_names, masterlist, xlsx_url, xlsx_sha)

    previous_path, previous = latest_snapshot()
    changes = None
    if not args.baseline:
        if previous is None:
            sys.stderr.write(
                "No previous snapshot found. Run with --baseline first.\n"
            )
            return 1
        changes = diff_snapshots(previous, snapshot)

    report = render_report(snapshot, changes, previous if not args.baseline else None)

    written = snapshot_path(snapshot["run_date"])
    written.write_text(json.dumps(snapshot, indent=2, ensure_ascii=False), encoding="utf-8")

    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    report_file = REPORT_DIR / f"livertox-{snapshot['run_date']}.md"
    counter = 2
    while report_file.exists():
        report_file = REPORT_DIR / f"livertox-{snapshot['run_date']}-{counter}.md"
        counter += 1
    report_file.write_text(report, encoding="utf-8")

    print(report)
    print(f"\nSnapshot: {written}")
    print(f"Report:   {report_file}")
    if previous_path and not args.baseline:
        print(f"Compared against: {previous_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
